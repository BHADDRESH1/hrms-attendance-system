from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from datetime import datetime, date, time, timedelta
from typing import List, Optional
import uuid

from app.modules.attendance.models import Attendance, AttendanceEdit
from app.modules.employees.models import Employee
from app.modules.attendance import schemas
from app.core.exceptions import HRMSException

# Corporate scheduling limits
CORPORATE_START_TIME = time(9, 0)
GRACE_PERIOD_MINUTES = 15

def _calculate_hours(clock_in: Optional[datetime], clock_out: Optional[datetime]) -> Optional[float]:
    if not clock_in or not clock_out:
        return None
    in_naive = clock_in.replace(tzinfo=None) if clock_in.tzinfo else clock_in
    out_naive = clock_out.replace(tzinfo=None) if clock_out.tzinfo else clock_out
    delta = out_naive - in_naive
    return round(delta.total_seconds() / 3600.0, 2)

async def punch_in(db: AsyncSession, employee: Employee, punch: schemas.AttendancePunchIn) -> Attendance:
    today = date.today()
    
    # Check duplicate punch
    stmt = select(Attendance).where(
        Attendance.employee_id == employee.id,
        Attendance.work_date == today
    )
    res = await db.execute(stmt)
    record = res.scalar_one_or_none()
    
    if record and record.clock_in is not None:
        raise HRMSException(message="Already clocked in for today.", status_code=400)
    
    now = datetime.now()
    status = "present"
    
    # Calculate late status
    start_limit = datetime.combine(today, CORPORATE_START_TIME) + timedelta(minutes=GRACE_PERIOD_MINUTES)
    if now > start_limit:
        status = "late"
        
    if not record:
        record = Attendance(
            employee_id=employee.id,
            work_date=today,
            clock_in=now,
            clock_in_ip=punch.clock_in_ip,
            clock_in_location=punch.clock_in_location,
            status=status
        )
        db.add(record)
    else:
        record.clock_in = now
        record.clock_in_ip = punch.clock_in_ip
        record.clock_in_location = punch.clock_in_location
        record.status = status
        db.add(record)
        
    await db.flush()
    await db.refresh(record)
    return record


async def punch_out(db: AsyncSession, employee: Employee, punch: schemas.AttendancePunchOut) -> Attendance:
    today = date.today()
    
    stmt = select(Attendance).where(
        Attendance.employee_id == employee.id,
        Attendance.work_date == today
    )
    res = await db.execute(stmt)
    record = res.scalar_one_or_none()
    
    if not record or record.clock_in is None:
        raise HRMSException(message="Cannot clock out without clocking in first.", status_code=400)
        
    if record.clock_out is not None:
        raise HRMSException(message="Already clocked out for today.", status_code=400)
        
    now = datetime.now()
    record.clock_out = now
    record.clock_out_ip = punch.clock_out_ip
    record.clock_out_location = punch.clock_out_location
    
    # Calculate duration
    hours = _calculate_hours(record.clock_in, record.clock_out)
    record.total_work_hours = hours
    
    if hours is None:
        hours = 0.0
        
    if hours < 5.0:
        record.status = "absent"
    elif hours < 7.0:
        record.status = "half_day"
    else:
        record.status = "present"
    
    db.add(record)
    await db.flush()
    await db.refresh(record)
    return record


async def get_attendance_history(
    db: AsyncSession,
    employee_id: uuid.UUID,
    start_date: date,
    end_date: date
) -> List[Attendance]:
    stmt = select(Attendance).where(
        Attendance.employee_id == employee_id,
        Attendance.work_date >= start_date,
        Attendance.work_date <= end_date
    ).order_by(Attendance.work_date.desc())
    
    res = await db.execute(stmt)
    return list(res.scalars().all())


async def get_daily_attendance(
    db: AsyncSession,
    work_date: date
) -> List[Attendance]:
    stmt = select(Attendance).where(Attendance.work_date == work_date).options(
        selectinload(Attendance.employee)
    )
    res = await db.execute(stmt)
    return list(res.scalars().all())


# --- Attendance Edit Services ---
async def request_attendance_edit(
    db: AsyncSession,
    employee: Employee,
    edit_in: schemas.AttendanceEditCreate
) -> AttendanceEdit:
    # Verify attendance record exists
    stmt = select(Attendance).where(Attendance.id == edit_in.attendance_id)
    res = await db.execute(stmt)
    attendance_rec = res.scalar_one_or_none()
    
    if not attendance_rec:
        raise HRMSException(message="Attendance record not found.", status_code=404)
        
    # Standard employee can only edit their own records
    if employee.user.role.name not in ["Super Admin", "Admin"] and attendance_rec.employee_id != employee.id:
        raise HRMSException(message="You do not have permission to request an edit for another employee's record.", status_code=403)

    # Prevent concurrent pending requests for the same record
    stmt = select(AttendanceEdit).where(
        AttendanceEdit.attendance_id == edit_in.attendance_id,
        AttendanceEdit.status == "pending"
    )
    res = await db.execute(stmt)
    if res.scalars().all():
        raise HRMSException(message="There is already a pending edit request for this attendance record.", status_code=400)

    edit_request = AttendanceEdit(
        attendance_id=edit_in.attendance_id,
        requested_by_id=employee.id,
        original_clock_in=attendance_rec.clock_in,
        original_clock_out=attendance_rec.clock_out,
        new_clock_in=edit_in.new_clock_in,
        new_clock_out=edit_in.new_clock_out,
        old_status=attendance_rec.status,
        new_status=attendance_rec.status,
        reason=edit_in.reason,
        status="pending"
    )
    
    db.add(edit_request)
    await db.flush()
    await db.refresh(edit_request)
    return edit_request


async def review_attendance_edit(
    db: AsyncSession,
    edit_id: uuid.UUID,
    approval: schemas.AttendanceEditApproval,
    reviewer: Employee
) -> AttendanceEdit:
    # Fetch edit request
    stmt = select(AttendanceEdit).where(AttendanceEdit.id == edit_id).options(
        selectinload(AttendanceEdit.attendance_record)
    )
    res = await db.execute(stmt)
    edit_req = res.scalar_one_or_none()
    
    if not edit_req:
        raise HRMSException(message="Attendance edit request not found.", status_code=404)
        
    if edit_req.status != "pending":
        raise HRMSException(message=f"Request has already been {edit_req.status}.", status_code=400)
        
    # Map status
    status_lower = approval.status.lower()
    if status_lower not in ["approved", "rejected"]:
        raise HRMSException(message="Status must be 'approved' or 'rejected'.", status_code=400)
        
    edit_req.status = status_lower
    edit_req.approved_by_id = reviewer.id
    
    if status_lower == "approved":
        # Modify original record
        attendance_rec = edit_req.attendance_record
        if edit_req.new_clock_in:
            attendance_rec.clock_in = edit_req.new_clock_in
        if edit_req.new_clock_out:
            attendance_rec.clock_out = edit_req.new_clock_out
            
        # Re-calculate hours
        hours = _calculate_hours(attendance_rec.clock_in, attendance_rec.clock_out)
        attendance_rec.total_work_hours = hours
            
        # Recalculate status
        if hours is not None:
            if hours < 5.0:
                attendance_rec.status = "absent"
            elif hours < 7.0:
                attendance_rec.status = "half_day"
            else:
                attendance_rec.status = "present"
        
        edit_req.new_status = attendance_rec.status
        db.add(attendance_rec)
        
    db.add(edit_req)
    await db.flush()
    await db.refresh(edit_req)
    return edit_req


async def list_pending_edits(db: AsyncSession) -> List[AttendanceEdit]:
    stmt = select(AttendanceEdit).where(AttendanceEdit.status == "pending").options(
        selectinload(AttendanceEdit.requested_by),
        selectinload(AttendanceEdit.attendance_record)
    )
    res = await db.execute(stmt)
    return list(res.scalars().all())


async def get_super_admin_analytics(
    db: AsyncSession,
    start_date: date,
    end_date: date
) -> dict:
    from app.modules.employees.models import Employee, User
    from sqlalchemy import func
    
    # 1. Total active employees
    emp_stmt = select(func.count(Employee.id)).join(User).where(User.is_active == True)
    emp_res = await db.execute(emp_stmt)
    total_employees = emp_res.scalar_one() or 0

    # 2. Number of working days (Mon-Fri)
    working_days = 0
    curr = start_date
    while curr <= end_date:
        if curr.weekday() < 5:  # Mon is 0, Fri is 4
            working_days += 1
        curr += timedelta(days=1)

    # 3. Sum status types and total productive hours
    stmt = select(
        func.count(Attendance.id).filter(Attendance.status == "present").label("present_count"),
        func.count(Attendance.id).filter(Attendance.status == "half_day").label("half_count"),
        func.count(Attendance.id).filter(Attendance.status == "on_leave").label("leave_count"),
        func.count(Attendance.id).filter(Attendance.status == "absent").label("lop_count"),
        func.count(Attendance.id).filter(Attendance.status == "permission").label("permission_count"),
        func.sum(Attendance.total_work_hours).label("productive_hours")
    ).where(
        Attendance.work_date >= start_date,
        Attendance.work_date <= end_date
    )
    res = await db.execute(stmt)
    row = res.first()

    present_days = row.present_count if row else 0
    half_days = row.half_count if row else 0
    leave_days = row.leave_count if row else 0
    lop_days = row.lop_count if row else 0
    permission_days = row.permission_count if row else 0
    actual_productive_hours = float(row.productive_hours or 0.0) if row else 0.0

    # Expected Hours = Total Employees x Working Days x 8
    expected_hours = float(total_employees * working_days * 8)
    
    # Lost Hours = Expected Hours - Actual Productive Hours
    lost_hours = max(0.0, expected_hours - actual_productive_hours)
    
    # Attendance Efficiency = (Actual Productive Hours / Expected Hours) x 100
    efficiency = (actual_productive_hours / expected_hours * 100) if expected_hours > 0 else 0.0

    return {
        "total_employees": total_employees,
        "working_days": working_days,
        "expected_hours": round(expected_hours, 2),
        "actual_productive_hours": round(actual_productive_hours, 2),
        "lost_hours": round(lost_hours, 2),
        "efficiency": round(efficiency, 2),
        "present_days": present_days,
        "half_days": half_days,
        "leave_days": leave_days,
        "lop_days": lop_days,
        "permission_days": permission_days
    }


async def update_attendance_admin(
    db: AsyncSession,
    record_id: uuid.UUID,
    update_in: schemas.AttendanceUpdateAdmin,
    admin_emp: Employee
) -> Attendance:
    stmt = select(Attendance).where(Attendance.id == record_id)
    res = await db.execute(stmt)
    record = res.scalar_one_or_none()
    
    if not record:
        raise HRMSException(message="Attendance record not found.", status_code=404)
        
    # Copy original values for audit logging
    original_in = record.clock_in
    original_out = record.clock_out
    original_status = record.status
    
    # Apply updates
    if update_in.clock_in is not None:
        record.clock_in = update_in.clock_in
    if update_in.clock_out is not None:
        record.clock_out = update_in.clock_out
        
    # Recalculate duration
    hours = _calculate_hours(record.clock_in, record.clock_out)
    record.total_work_hours = hours
        
    # Recalculate status and value
    if update_in.status is not None:
        record.status = update_in.status
    else:
        if hours is not None:
            if hours < 5.0:
                record.status = "absent"
            elif hours < 7.0:
                record.status = "half_day"
            else:
                record.status = "present"
                
    # Create approved edit request log for audit history
    edit_log = AttendanceEdit(
        attendance_id=record.id,
        requested_by_id=admin_emp.id,
        approved_by_id=admin_emp.id,
        original_clock_in=original_in,
        original_clock_out=original_out,
        new_clock_in=record.clock_in,
        new_clock_out=record.clock_out,
        old_status=original_status,
        new_status=record.status,
        reason=f"[Super Admin Override]: {update_in.reason}",
        status="approved"
    )
    
    db.add(record)
    db.add(edit_log)
    await db.flush()
    await db.refresh(record)
    return record


async def get_attendance_audit_logs(
    db: AsyncSession,
    employee_id: Optional[uuid.UUID] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 10
) -> tuple[List[AttendanceEdit], int]:
    from app.modules.employees.models import Employee
    from sqlalchemy import func
    
    # Base query for rows
    stmt = select(AttendanceEdit).join(AttendanceEdit.attendance_record).join(Attendance.employee).options(
        selectinload(AttendanceEdit.attendance_record).selectinload(Attendance.employee),
        selectinload(AttendanceEdit.requested_by),
        selectinload(AttendanceEdit.approved_by)
    ).where(AttendanceEdit.status == "approved")
    
    # Base query for counts
    count_stmt = select(func.count(AttendanceEdit.id)).join(AttendanceEdit.attendance_record).join(Attendance.employee).where(AttendanceEdit.status == "approved")
    
    # Apply filters
    if employee_id:
        stmt = stmt.where(Attendance.employee_id == employee_id)
        count_stmt = count_stmt.where(Attendance.employee_id == employee_id)
        
    if start_date:
        stmt = stmt.where(Attendance.work_date >= start_date)
        count_stmt = count_stmt.where(Attendance.work_date >= start_date)
        
    if end_date:
        stmt = stmt.where(Attendance.work_date <= end_date)
        count_stmt = count_stmt.where(Attendance.work_date <= end_date)
        
    if search:
        search_term = f"%{search}%"
        search_filter = (
            (Employee.first_name.ilike(search_term)) |
            (Employee.last_name.ilike(search_term)) |
            (Employee.employee_id_code.ilike(search_term)) |
            (AttendanceEdit.reason.ilike(search_term))
        )
        stmt = stmt.where(search_filter)
        count_stmt = count_stmt.where(search_filter)
        
    # Get total count
    count_res = await db.execute(count_stmt)
    total_count = count_res.scalar_one() or 0
    
    # Get rows with pagination offset
    offset = (page - 1) * limit
    stmt = stmt.order_by(AttendanceEdit.created_at.desc()).offset(offset).limit(limit)
    res = await db.execute(stmt)
    rows = list(res.scalars().all())
    
    return rows, total_count


async def query_attendance_for_export(
    db: AsyncSession,
    department_id: Optional[uuid.UUID] = None,
    employee_id: Optional[uuid.UUID] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None
) -> List[Attendance]:
    from app.modules.employees.models import Employee
    stmt = select(Attendance).join(Attendance.employee).options(
        selectinload(Attendance.employee).selectinload(Employee.department)
    )
    
    if department_id:
        stmt = stmt.where(Employee.department_id == department_id)
    if employee_id:
        stmt = stmt.where(Attendance.employee_id == employee_id)
    if start_date:
        stmt = stmt.where(Attendance.work_date >= start_date)
    if end_date:
        stmt = stmt.where(Attendance.work_date <= end_date)
    if status:
        stmt = stmt.where(Attendance.status == status)
        
    stmt = stmt.order_by(Employee.employee_id_code, Attendance.work_date.asc())
    res = await db.execute(stmt)
    return list(res.scalars().all())


def generate_attendance_csv(records: List[Attendance]) -> str:
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(["Employee ID", "Employee Name", "Department", "Work Date", "Clock In", "Clock Out", "Hours Worked", "Status"])
    
    for rec in records:
        dept_name = rec.employee.department.name if rec.employee.department else "N/A"
        clock_in_str = rec.clock_in.strftime("%Y-%m-%d %H:%M:%S") if rec.clock_in else "N/A"
        clock_out_str = rec.clock_out.strftime("%Y-%m-%d %H:%M:%S") if rec.clock_out else "N/A"
        writer.writerow([
            rec.employee.employee_id_code,
            f"{rec.employee.first_name} {rec.employee.last_name}",
            dept_name,
            str(rec.work_date),
            clock_in_str,
            clock_out_str,
            rec.total_work_hours if rec.total_work_hours is not None else 0.0,
            rec.status
        ])
        
    return output.getvalue()


def generate_attendance_xlsx(records: List[Attendance]) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Workforce Attendance Logs"
    
    # Grid Styling details
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Headers
    headers = ["Employee ID", "Employee Name", "Department", "Work Date", "Clock In", "Clock Out", "Hours Worked", "Status"]
    ws.append(headers)
    
    # Header format
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align if col not in [2, 3] else left_align
        
    # Data rows
    for rec in records:
        dept_name = rec.employee.department.name if rec.employee.department else "N/A"
        clock_in_str = rec.clock_in.strftime("%Y-%m-%d %H:%M:%S") if rec.clock_in else "N/A"
        clock_out_str = rec.clock_out.strftime("%Y-%m-%d %H:%M:%S") if rec.clock_out else "N/A"
        ws.append([
            rec.employee.employee_id_code,
            f"{rec.employee.first_name} {rec.employee.last_name}",
            dept_name,
            str(rec.work_date),
            clock_in_str,
            clock_out_str,
            rec.total_work_hours if rec.total_work_hours is not None else 0.0,
            rec.status.capitalize()
        ])
        
    # Auto column spacing sizes
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        for cell in col:
            if cell.row > 1:
                cell.font = data_font
                
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_attendance_pdf(records: List[Attendance]) -> bytes:
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=15
    )
    
    story.append(Paragraph("Workforce Attendance Roster Report", title_style))
    story.append(Spacer(1, 10))
    
    # Table data
    data = [["Emp ID", "Name", "Date", "Clock In", "Clock Out", "Hours", "Status"]]
    for rec in records:
        clock_in_str = rec.clock_in.strftime("%I:%M %p") if rec.clock_in else "--:--"
        clock_out_str = rec.clock_out.strftime("%I:%M %p") if rec.clock_out else "--:--"
        data.append([
            rec.employee.employee_id_code,
            f"{rec.employee.first_name} {rec.employee.last_name}",
            str(rec.work_date),
            clock_in_str,
            clock_out_str,
            str(rec.total_work_hours) if rec.total_work_hours is not None else "0.0",
            rec.status.capitalize()
        ])
        
    table = Table(data, colWidths=[60, 110, 80, 75, 75, 50, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0284c7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#334155')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
    ]))
    
    story.append(table)
    doc.build(story)
    return buffer.getvalue()
