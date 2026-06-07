import uuid
from datetime import datetime, date, time, timedelta, timezone
from typing import List, Optional, Dict, Any
import calendar
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import func, text

from app.modules.attendance.models import Attendance
from app.modules.employees.models import Employee, Department, User

def _format_time(dt: Optional[datetime]) -> str:
    if not dt:
        return "-"
    # Convert UTC to local format (usually naive or format directly)
    # Since we store as timezone-aware, let's format it to "HH:MM AM/PM"
    return dt.strftime("%I:%M %p")

async def get_company_summary(db: AsyncSession, month: int, year: int) -> Dict[str, Any]:
    # 1. Total active employees
    emp_stmt = select(func.count(Employee.id)).join(User).where(User.is_active == True)
    emp_res = await db.execute(emp_stmt)
    total_employees = emp_res.scalar_one() or 0

    today = date.today()

    # 2. Present today (present or late status today)
    present_stmt = select(func.count(Attendance.id)).where(
        Attendance.work_date == today,
        Attendance.status.in_(["present", "late"])
    )
    present_res = await db.execute(present_stmt)
    present_today = present_res.scalar_one() or 0

    # 3. Leave today
    leave_stmt = select(func.count(Attendance.id)).where(
        Attendance.work_date == today,
        Attendance.status.in_(["leave", "on_leave"])
    )
    leave_res = await db.execute(leave_stmt)
    leave_today = leave_res.scalar_one() or 0

    # 4. Absent today
    absent_today = max(0, total_employees - present_today - leave_today)

    # 5. Total monthly hours
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)
    
    hours_stmt = select(func.sum(Attendance.total_work_hours)).where(
        Attendance.work_date >= start_date,
        Attendance.work_date <= end_date
    )
    hours_res = await db.execute(hours_stmt)
    total_monthly_hours = float(hours_res.scalar_one() or 0.0)

    return {
        "total_employees": total_employees,
        "present_today": present_today,
        "absent_today": absent_today,
        "leave_today": leave_today,
        "total_monthly_hours": round(total_monthly_hours, 2)
    }

async def get_employee_monthly_analytics(
    db: AsyncSession,
    employee_id: uuid.UUID,
    month: int,
    year: int
) -> Optional[Dict[str, Any]]:
    # Fetch employee details
    emp_stmt = select(Employee).where(Employee.id == employee_id).options(
        selectinload(Employee.user).selectinload(User.role),
        selectinload(Employee.department)
    )
    emp_res = await db.execute(emp_stmt)
    employee = emp_res.scalar_one_or_none()
    if not employee:
        return None

    joined_date = employee.joined_date or date(2026, 1, 1)

    # Calculate days in month
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)

    # Fetch attendance records
    att_stmt = select(Attendance).where(
        Attendance.employee_id == employee_id,
        Attendance.work_date >= start_date,
        Attendance.work_date <= end_date
    )
    att_res = await db.execute(att_stmt)
    records = {r.work_date: r for r in att_res.scalars().all()}

    timeline = []
    working_days = 0
    present_days = 0
    leave_days = 0
    absent_days = 0
    holiday_days = 0
    total_hours = 0.0

    # Build daily timeline
    for d in range(1, num_days + 1):
        curr_date = date(year, month, d)
        day_name = curr_date.strftime("%A")
        is_weekend = curr_date.weekday() >= 5 # 5=Saturday, 6=Sunday

        if curr_date < joined_date:
            timeline.append({
                "date": curr_date.isoformat(),
                "day": day_name,
                "status": "N/A",
                "check_in": "-",
                "check_out": "-",
                "hours_worked": "-"
            })
        else:
            rec = records.get(curr_date)
            # Count working days (weekdays on/after joined date)
            if not is_weekend:
                working_days += 1

            if rec:
                # Map status correctly for display
                status_raw = rec.status.lower()
                if status_raw in ["present", "late"]:
                    status = "Present"
                    present_days += 1
                elif status_raw == "half_day":
                    status = "Half Day"
                    present_days += 0.5
                elif status_raw in ["leave", "on_leave"]:
                    status = "Leave"
                    leave_days += 1
                elif status_raw in ["absent", "lop"]:
                    status = "Absent"
                    absent_days += 1
                else:
                    status = rec.status.capitalize()
                
                check_in = _format_time(rec.clock_in)
                check_out = _format_time(rec.clock_out)
                
                # Check for N/A display conditions
                if status in ["Leave", "Holiday", "N/A"]:
                    hours_val = "-"
                    check_in = "-"
                    check_out = "-"
                else:
                    hours_val = float(rec.total_work_hours) if rec.total_work_hours is not None else 0.0
                    total_hours += hours_val if isinstance(hours_val, float) else 0.0
                    hours_val = f"{hours_val:.2f}"
                
                timeline.append({
                    "date": curr_date.isoformat(),
                    "day": day_name,
                    "status": status,
                    "check_in": check_in,
                    "check_out": check_out,
                    "hours_worked": hours_val
                })
            else:
                if is_weekend:
                    status = "Holiday"
                    holiday_days += 1
                else:
                    status = "Absent"
                    absent_days += 1

                timeline.append({
                    "date": curr_date.isoformat(),
                    "day": day_name,
                    "status": status,
                    "check_in": "-",
                    "check_out": "-",
                    "hours_worked": "-"
                })

    # Calculations
    attendance_pct = (present_days / working_days * 100) if working_days > 0 else 0.0
    avg_hours = (total_hours / present_days) if present_days > 0 else 0.0

    # Weekly Productivity data (Total hours per calendar week)
    # We group days by calendar week (1 to 5)
    weekly_hours = [0.0] * 6
    for t in timeline:
        if t["hours_worked"] != "-":
            d_obj = date.fromisoformat(t["date"])
            w_idx = (d_obj.day - 1) // 7
            weekly_hours[w_idx] += float(t["hours_worked"])

    weekly_productivity = [
        {"week": f"Week {i+1}", "hours": round(weekly_hours[i], 2)} for i in range(5) if i < (num_days + 6) // 7
    ]

    # Annual Monthly Working Hours Graph
    # We fetch total hours worked per month in the selected year
    monthly_hours_data = []
    for m in range(1, 12 + 1):
        m_start = date(year, m, 1)
        m_end = date(year, m, calendar.monthrange(year, m)[1])
        m_stmt = select(func.sum(Attendance.total_work_hours)).where(
            Attendance.employee_id == employee_id,
            Attendance.work_date >= m_start,
            Attendance.work_date <= m_end
        )
        m_res = await db.execute(m_stmt)
        m_hours = float(m_res.scalar_one() or 0.0)
        monthly_hours_data.append({
            "month": calendar.month_abbr[m],
            "hours": round(m_hours, 2)
        })

    # Pie Chart distribution counts
    # We map status types count across all timeline days
    pie_counts = {
        "Present": 0,
        "Absent": 0,
        "Leave": 0,
        "Holiday": 0
    }
    for t in timeline:
        st = t["status"]
        if st in ["Present", "Half Day"]:
            pie_counts["Present"] += 1
        elif st == "Leave":
            pie_counts["Leave"] += 1
        elif st == "Holiday":
            pie_counts["Holiday"] += 1
        elif st == "Absent":
            pie_counts["Absent"] += 1

    total_pie_days = sum(pie_counts.values())
    pie_chart_data = [
        {"name": k, "value": round(v / total_pie_days * 100, 1) if total_pie_days > 0 else 0.0}
        for k, v in pie_counts.items()
    ]

    # Manager display name
    manager_name = employee.reporting_manager or "None"

    return {
        "employee": {
            "id": employee.id,
            "employee_id_code": employee.employee_id_code,
            "name": employee.full_name,
            "department": employee.department.name if employee.department else "N/A",
            "designation": employee.designation or "Staff",
            "manager": manager_name,
            "joined_date": joined_date.isoformat()
        },
        "kpis": {
            "working_days": working_days,
            "present_days": present_days,
            "leave_days": leave_days,
            "absent_days": absent_days,
            "attendance_percentage": round(attendance_pct, 1),
            "total_working_hours": round(total_hours, 2),
            "average_working_hours": round(avg_hours, 2)
        },
        "timeline": timeline,
        "weekly_productivity": weekly_productivity,
        "monthly_hours_trend": monthly_hours_data,
        "pie_chart_distribution": pie_chart_data
    }

async def get_departments_analytics(db: AsyncSession, month: int, year: int) -> List[Dict[str, Any]]:
    # Fetch all departments
    dept_stmt = select(Department)
    dept_res = await db.execute(dept_stmt)
    departments = dept_res.scalars().all()
    
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)

    analytics = []

    # Map department employees
    for dept in departments:
        # Fetch active employees
        emp_stmt = select(Employee).join(User).where(Employee.department_id == dept.id, User.is_active == True)
        emp_res = await db.execute(emp_stmt)
        employees = emp_res.scalars().all()
        emp_ids = [e.id for e in employees]

        total_hours = 0.0
        if emp_ids:
            hours_stmt = select(func.sum(Attendance.total_work_hours)).where(
                Attendance.employee_id.in_(emp_ids),
                Attendance.work_date >= start_date,
                Attendance.work_date <= end_date
            )
            hours_res = await db.execute(hours_stmt)
            total_hours = float(hours_res.scalar_one() or 0.0)

        emp_count = len(employees)
        avg_hours = (total_hours / emp_count) if emp_count > 0 else 0.0

        analytics.append({
            "department_name": dept.name,
            "total_hours": round(total_hours, 2),
            "average_hours": round(avg_hours, 2),
            "employee_count": emp_count
        })

    # "Other Departments" group for employees without assigned department
    other_emp_stmt = select(Employee).join(User).where(Employee.department_id == None, User.is_active == True)
    other_emp_res = await db.execute(other_emp_stmt)
    other_employees = other_emp_res.scalars().all()
    other_emp_ids = [e.id for e in other_employees]

    other_total_hours = 0.0
    if other_emp_ids:
        hours_stmt = select(func.sum(Attendance.total_work_hours)).where(
            Attendance.employee_id.in_(other_emp_ids),
            Attendance.work_date >= start_date,
            Attendance.work_date <= end_date
        )
        hours_res = await db.execute(hours_stmt)
        other_total_hours = float(hours_res.scalar_one() or 0.0)

    other_count = len(other_employees)
    other_avg_hours = (other_total_hours / other_count) if other_count > 0 else 0.0

    if other_count > 0 or other_total_hours > 0:
        analytics.append({
            "department_name": "Other Departments",
            "total_hours": round(other_total_hours, 2),
            "average_hours": round(other_avg_hours, 2),
            "employee_count": other_count
        })

    return analytics


def generate_employee_monthly_xlsx(analytics: Dict[str, Any]) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    # Style definitions
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    section_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    title_font = Font(name="Calibri", size=14, bold=True, color="0284C7")
    section_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True)
    data_font = Font(name="Calibri", size=10)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    emp_info = analytics["employee"]
    kpis = analytics["kpis"]
    
    ws.cell(row=1, column=1, value="HRMS Attendance Report").font = title_font
    ws.cell(row=2, column=1, value=f"Monthly Summary for {emp_info['name']}").font = Font(name="Calibri", size=10, italic=True)
    
    ws.cell(row=4, column=1, value="Employee Information").font = section_font
    ws.merge_cells("A4:F4")
    ws.cell(row=4, column=1).fill = section_fill
    
    emp_details = [
        ("Name", emp_info["name"], "Employee ID", emp_info["employee_id_code"]),
        ("Department", emp_info["department"], "Designation", emp_info["designation"]),
        ("Manager", emp_info["manager"], "Joined Date", emp_info["joined_date"])
    ]
    
    row_idx = 5
    for label1, val1, label2, val2 in emp_details:
        ws.cell(row=row_idx, column=1, value=label1).font = bold_font
        ws.cell(row=row_idx, column=2, value=val1).font = data_font
        ws.cell(row=row_idx, column=4, value=label2).font = bold_font
        ws.cell(row=row_idx, column=5, value=val2).font = data_font
        row_idx += 1
        
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Key Performance Indicators (KPIs)").font = section_font
    ws.merge_cells(f"A{row_idx}:F{row_idx}")
    ws.cell(row=row_idx, column=1).fill = section_fill
    
    row_idx += 1
    kpi_items = [
        ("Working Days", f"{kpis['working_days']} Days"),
        ("Present Days", f"{kpis['present_days']} Days"),
        ("Leave Days", f"{kpis['leave_days']} Days"),
        ("Absent Days", f"{kpis['absent_days']} Days"),
        ("Attendance %", f"{kpis['attendance_percentage']}%"),
        ("Total Hours", f"{kpis['total_working_hours']} hrs")
    ]
    
    for col_idx, (kpi_name, kpi_val) in enumerate(kpi_items, 1):
        ws.cell(row=row_idx, column=col_idx, value=kpi_name).font = bold_font
        ws.cell(row=row_idx, column=col_idx).alignment = center_align
        ws.cell(row=row_idx+1, column=col_idx, value=kpi_val).font = data_font
        ws.cell(row=row_idx+1, column=col_idx).alignment = center_align
        
    row_idx += 3
    ws.cell(row=row_idx, column=1, value="Detailed Daily Attendance Logs").font = section_font
    ws.merge_cells(f"A{row_idx}:F{row_idx}")
    ws.cell(row=row_idx, column=1).fill = section_fill
    
    row_idx += 1
    headers = ["Date", "Day", "Status", "Check In", "Check Out", "Hours Worked"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    for day in analytics["timeline"]:
        row_idx += 1
        row_data = [day["date"], day["day"], day["status"], day["check_in"], day["check_out"], day["hours_worked"]]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx == 3:
                status_lower = str(val).lower()
                if status_lower in ["present", "present days", "late"]:
                    cell.font = Font(name="Calibri", size=10, color="10B981", bold=True)
                elif status_lower in ["absent", "absent days"]:
                    cell.font = Font(name="Calibri", size=10, color="EF4444", bold=True)
                elif status_lower in ["leave", "leave days"]:
                    cell.font = Font(name="Calibri", size=10, color="3B82F6", bold=True)
                elif status_lower == "holiday":
                    cell.font = Font(name="Calibri", size=10, color="F59E0B", bold=True)
                    
            if col_idx in [1, 2, 4, 5]:
                cell.alignment = center_align
            elif col_idx == 6:
                cell.alignment = right_align
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_employee_monthly_csv(analytics: Dict[str, Any]) -> str:
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    emp_info = analytics["employee"]
    kpis = analytics["kpis"]
    
    writer.writerow(["HRMS Attendance Report"])
    writer.writerow([])
    
    writer.writerow(["Employee Details"])
    writer.writerow(["Name", emp_info["name"], "Employee ID", emp_info["employee_id_code"]])
    writer.writerow(["Department", emp_info["department"], "Designation", emp_info["designation"]])
    writer.writerow(["Manager", emp_info["manager"], "Joined Date", emp_info["joined_date"]])
    writer.writerow([])
    
    writer.writerow(["Attendance KPIs"])
    writer.writerow(["Working Days", f"{kpis['working_days']} Days"])
    writer.writerow(["Present Days", f"{kpis['present_days']} Days"])
    writer.writerow(["Leave Days", f"{kpis['leave_days']} Days"])
    writer.writerow(["Absent Days", f"{kpis['absent_days']} Days"])
    writer.writerow(["Attendance Percentage", f"{kpis['attendance_percentage']}%"])
    writer.writerow(["Total Working Hours", f"{kpis['total_working_hours']} Hours"])
    writer.writerow([])
    
    writer.writerow(["Daily Logs"])
    writer.writerow(["Date", "Day", "Status", "Check In", "Check Out", "Hours Worked"])
    for day in analytics["timeline"]:
        writer.writerow([
            day["date"],
            day["day"],
            day["status"],
            day["check_in"],
            day["check_out"],
            day["hours_worked"]
        ])
        
    return output.getvalue()
